import pandas as pd
import os

from openpyxl.styles import (
    Font,
    PatternFill
)

from openpyxl.utils import (
    get_column_letter
)


def flatten_metadata(record):

    row = {}

    row["TC ID"] = record.get(
        "tc_id",
        ""
    )

    row["Action"] = record.get(
        "action",
        ""
    )

    row["Description"] = record.get(
        "description",
        ""
    )

    metadata = record.get(
        "metadata",
        {}
    )

    row["Module"] = metadata.get(
        "module",
        ""
    )

    row["Feature"] = metadata.get(
        "feature",
        ""
    )

    row["Title"] = metadata.get(
        "title",
        ""
    )

    row["Priority"] = metadata.get(
        "priority",
        ""
    )

    row["Severity"] = metadata.get(
        "severity",
        ""
    )

    row["Document"] = (
        record.get(
            "document",
            ""
        )[:500]
    )

    return row


def format_sheet(sheet):

    header_fill = PatternFill(
        start_color="1F4E78",
        end_color="1F4E78",
        fill_type="solid"
    )

    header_font = Font(
        color="FFFFFF",
        bold=True
    )

    for cell in sheet[1]:

        cell.fill = header_fill
        cell.font = header_font

    for column in sheet.columns:

        max_length = 0

        column_letter = (
            get_column_letter(
                column[0].column
            )
        )

        for cell in column:

            try:

                if len(
                    str(cell.value)
                ) > max_length:

                    max_length = len(
                        str(cell.value)
                    )

            except:
                pass

        adjusted_width = min(
            max_length + 5,
            50
        )

        sheet.column_dimensions[
            column_letter
        ].width = adjusted_width

    sheet.auto_filter.ref = (
        sheet.dimensions
    )


def export_impact_analysis_v3(
    analysis_result
):

    os.makedirs(
        "output",
        exist_ok=True
    )

    file_path = (
        "output/impact_analysis_report_v3.xlsx"
    )

    # ---------------------------------
    # Handle Both Response Formats
    # ---------------------------------

    if (
        "impact_analysis"
        in analysis_result
    ):

        data = analysis_result[
            "impact_analysis"
        ]

    else:

        data = analysis_result

    impacted = data.get(
        "impacted_test_cases",
        []
    )

    modified = data.get(
        "modified_test_cases",
        []
    )

    new_cases = data.get(
        "new_test_cases",
        []
    )

    coverage_gaps = data.get(
        "coverage_gaps",
        []
    )

    # ---------------------------------
    # Debugging
    # ---------------------------------

    print(
        "\n========== EXPORT DEBUG =========="
    )

    print(
        "Impacted:",
        len(impacted)
    )

    print(
        "Modified:",
        len(modified)
    )

    print(
        "New:",
        len(new_cases)
    )

    print(
        "Coverage:",
        len(coverage_gaps)
    )

    print(
        "Sample Impacted:",
        impacted[:1]
    )

    print(
        "Sample Modified:",
        modified[:1]
    )

    print(
        "==================================\n"
    )

    # ---------------------------------
    # Impacted Test Cases
    # ---------------------------------

    impacted_df = pd.DataFrame([
        {
            "TC ID":
                tc.get(
                    "tc_id",
                    ""
                ),

            "Reason":
                tc.get(
                    "reason",
                    ""
                )
        }

        for tc in impacted
    ])

    # ---------------------------------
    # Modified Test Cases
    # ---------------------------------

    modified_df = pd.DataFrame(
        [
            flatten_metadata(x)
            for x in modified
        ]
    )

    # ---------------------------------
    # New Test Cases
    # ---------------------------------

    new_df = pd.DataFrame(
        [
            flatten_metadata(x)
            for x in new_cases
        ]
    )

    # ---------------------------------
    # Coverage Gaps
    # ---------------------------------

    gaps_df = pd.DataFrame(
        coverage_gaps,
        columns=[
            "Coverage Gap"
        ]
    )

    # ---------------------------------
    # Write Excel
    # ---------------------------------

    with pd.ExcelWriter(
        file_path,
        engine="openpyxl"
    ) as writer:

        impacted_df.to_excel(
            writer,
            sheet_name=
            "Impacted_Test_Cases",
            index=False
        )

        modified_df.to_excel(
            writer,
            sheet_name=
            "Modified_Test_Cases",
            index=False
        )

        new_df.to_excel(
            writer,
            sheet_name=
            "New_Test_Cases",
            index=False
        )

        gaps_df.to_excel(
            writer,
            sheet_name=
            "Coverage_Gaps",
            index=False
        )

        workbook = writer.book

        for sheet_name in (
            workbook.sheetnames
        ):

            sheet = workbook[
                sheet_name
            ]

            format_sheet(
                sheet
            )

    print(
        f"\nProfessional Impact Report Saved: {file_path}\n"
    )

    return file_path